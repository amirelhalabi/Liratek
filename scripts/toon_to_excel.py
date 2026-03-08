#!/usr/bin/env python3
"""
Convert a .toon file to an Excel (.xlsx) spreadsheet.

Usage:
    python scripts/toon_to_excel.py                              # uses default file
    python scripts/toon_to_excel.py path/to/file.toon            # specify input
    python scripts/toon_to_excel.py path/to/file.toon -o out.xlsx  # specify output
"""

import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def parse_toon(filepath: Path) -> tuple[list[str], list[list[str]]]:
    """
    Parse a .toon file and return (headers, rows).

    Format:
        items[COUNT,]{field1,field2,...}:
          category,name,code,price,cost,supplier,unit
          ...
    """
    text = filepath.read_text(encoding="utf-8")
    lines = text.strip().splitlines()

    if not lines:
        print("Error: empty .toon file", file=sys.stderr)
        sys.exit(1)

    # Parse header line: items[4010,]{category,name,code,price,cost,supplier,unit}:
    header_match = re.match(r"items\[\d+,?\]\{(.+)\}:", lines[0])
    if not header_match:
        print(f"Error: invalid .toon header: {lines[0]}", file=sys.stderr)
        sys.exit(1)

    headers = [h.strip() for h in header_match.group(1).split(",")]

    rows = []
    for line in lines[1:]:
        line = line.strip()
        if not line:
            continue

        # Handle quoted fields (CSV-style) and commas within quotes
        fields = []
        current = ""
        in_quotes = False
        i = 0
        while i < len(line):
            ch = line[i]
            if ch == '"' and not in_quotes:
                in_quotes = True
                i += 1
                continue
            elif ch == '"' and in_quotes:
                # Check for escaped quote ""
                if i + 1 < len(line) and line[i + 1] == '"':
                    current += '"'
                    i += 2
                    continue
                in_quotes = False
                i += 1
                continue
            elif ch == "," and not in_quotes:
                fields.append(current.strip())
                current = ""
                i += 1
                continue
            else:
                current += ch
                i += 1
        fields.append(current.strip())

        rows.append(fields)

    return headers, rows


def write_excel(headers: list[str], rows: list[list[str]], output: Path):
    """Write parsed .toon data to a styled Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"

    # -- Header labels (prettified) --
    pretty = {
        "category": "Category",
        "name": "Product Name",
        "code": "Barcode",
        "price": "Price ($)",
        "cost": "Cost ($)",
        "supplier": "Supplier",
        "unit": "Unit",
    }
    header_labels = [pretty.get(h, h.title()) for h in headers]

    # -- Styles --
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    currency_fmt = '#,##0.00'
    int_fmt = '0'

    # Numeric column indices (0-based, relative to headers)
    numeric_cols = set()
    int_cols = set()
    for i, h in enumerate(headers):
        if h in ("price", "cost"):
            numeric_cols.add(i)
        elif h == "unit":
            int_cols.add(i)

    # -- Write header row --
    for col_idx, label in enumerate(header_labels, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # -- Write data rows --
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    for row_idx, row_data in enumerate(rows, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data):
            col_1based = col_idx + 1
            if col_1based > len(headers):
                break

            # Convert numeric fields
            cell_value = value
            if col_idx in numeric_cols:
                if value and value != "null":
                    try:
                        cell_value = float(value)
                    except ValueError:
                        pass
                else:
                    cell_value = None
            elif col_idx in int_cols:
                if value and value != "null":
                    try:
                        cell_value = int(float(value))
                    except ValueError:
                        pass
                else:
                    cell_value = 0

            # Replace "null" string with empty
            if cell_value == "null":
                cell_value = None

            cell = ws.cell(row=row_idx, column=col_1based, value=cell_value)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)

            if col_idx in numeric_cols and isinstance(cell_value, (int, float)):
                cell.number_format = currency_fmt
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in int_cols:
                cell.number_format = int_fmt
                cell.alignment = Alignment(horizontal="center")

            if fill:
                cell.fill = fill

    # -- Auto-fit column widths --
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(ws.cell(row=1, column=col_idx).value))
        for row_idx in range(2, min(len(rows) + 2, 200)):  # sample first 200 rows
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        # Cap width and add padding
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

    # -- Freeze header row --
    ws.freeze_panes = "A2"

    # -- Auto-filter --
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    wb.save(output)
    print(f"Wrote {len(rows)} items to {output}")


def main():
    # Default input file
    default_input = Path(__file__).resolve().parent.parent / "item.2026-03-02 (1).toon"
    toon_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_input

    # Output path
    if "-o" in sys.argv:
        out_path = Path(sys.argv[sys.argv.index("-o") + 1])
    else:
        out_path = toon_path.with_suffix(".xlsx")

    if not toon_path.exists():
        print(f"Error: file not found: {toon_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading: {toon_path}")
    headers, rows = parse_toon(toon_path)
    print(f"Parsed {len(rows)} items with columns: {', '.join(headers)}")
    write_excel(headers, rows, out_path)


if __name__ == "__main__":
    main()
